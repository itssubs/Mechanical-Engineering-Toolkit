import pandas as pd
from material import Material
class Materialdatabase:
    def __init__(self, filepath = None):
        if filepath:
            self.load(filepath)
        else:
            self._df = pd.DataFrame(columns = ['Material', 'Yield Strength', 'Density', 'E_GPa', 'Brinell', 'Poissons'])

    def load(self, filepath):
        self._df = pd.read_csv(filepath)
    
    def save(self, filepath):
        self._df.to_csv(filepath, index = False)

    def add(self, material):
        new_row = {'Material': material.name, 'Yield Strength': material.yield_mpa, 'Density': material.density, 'E_GPa' : material.E_GPa, 'Brinell': material.brinell, 'Poissons': material.poissons}
        self._df = pd.concat([self._df, pd.DataFrame([new_row])], ignore_index = True)
    
    def get(self, name):
        match = self._df[self._df["Material"] == name]

        if match.empty:
            raise KeyError(f"Material {name} not found in the directory")
        
        row = match.iloc[0]
        return Material(row["Material"], float(row["Yield Strength"]), float(row["Density"]), float(row["E_GPa"]), float(row['Brinell']), float(row['Poissons']))
    
    def filter_by_strength(self, min_mpa):
        match = self._df[self._df['Yield Strength'] >= min_mpa]

        if match.empty:
            return []
        
        material_objects = []
        for _, row in match.iterrows():
            mat_obj = Material(name = row['Material'], yield_mpa = float(row['Yield Strength']), density = float(row['Density']), E_GPa = float(row['E_GPa']), brinell = float(row['Brinell']), poissons = float(row['Poissons']))
            material_objects.append(mat_obj)

        return material_objects
    
    def best_strength_to_weight(self):
        strength_to_weight = self._df["Yield Strength"] / self._df["Density"]

        max_index = strength_to_weight.idxmax()
        best_row = self._df.loc[max_index]

        return Material(
            name=best_row["Material"],
            yield_mpa=float(best_row["Yield Strength"]),
            density=float(best_row["Density"]),
            E_GPa=float(best_row["E_GPa"]),
            brinell = float(best_row["Brinell"]),
            poissons = float(best_row["Poissons"]),
        )
    
    def to_excel(self, filepath):
        with pd.ExcelWriter(filepath) as writer:
            self._df.to_excel(writer, sheet_name = 'Materials', index = False)
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = load_workbook(filepath)
        for ws in wb.worksheets:
            ws.freeze_panes = 'A2'
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    if cell.value:
                        max_length = max(max_length , len(str(cell.value)))
                ws.column_dimensions[column_letter].width = max_length + 2

        #Header formatting
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.font = Font(bold = True)
                cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
                cell.fill = PatternFill(fill_type = 'solid', start_color = "00FFFB")
        
        wb.save(filepath)
